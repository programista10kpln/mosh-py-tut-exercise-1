import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    corrected_price_title = sheet.cell(1, 4)
    corrected_price_title.value = 'corrected_price'

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.number_format = '#,##$0.00'
        corrected_price_cell.value = corrected_price

    sheet._charts.clear()

    values = Reference(sheet, min_row=1, max_row=sheet.max_row, min_col=3, max_col=4)

    chart = BarChart()
    chart.add_data(values, titles_from_data=True)
    chart.title = 'Price corrections chart'
    chart.y_axis.title = 'Prices'
    sheet.add_chart(chart, 'a6')

    assert isinstance(filename, object)
    wb.save(filename)


process_workbook('transactions.xlsx')
